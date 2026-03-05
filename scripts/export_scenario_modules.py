"""
Export Scenario-Module overview from the FinOps Module Repository Excel file.

Joins: VBD -> Map-VBD-Scenario -> VBD Scenarios -> Map-Scenario-Module -> VBD Module
Output columns: VBD Title, Scenario Title, Module Id, Module Title, Suggested Order, Module Type,
               Module Lifecycle, Module first release date, Module last modified

Optionally generates a Markdown changelog of modules changed in the last 30 days (enabled by default).

Version: 1.5.0 (2026-03-05 19:09)

Usage:
    python export_scenario_modules.py
    python export_scenario_modules.py --input <path_to_xlsx> --output <output_folder>
    python export_scenario_modules.py --no-changelog
    python export_scenario_modules.py --days 60
"""

__version__ = "1.5.0 (2026-03-05 19:09)"

import argparse
import csv
import datetime
import os
import sys
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_EXCEL_FILE = os.path.join(SCRIPT_DIR, "..", "source", "FinOpsCost-Module-Repository.xlsx")
OUTPUT_FILENAME = "scenario_modules.csv"
CHANGELOG_FILENAME = "module_changes.md"


def to_date(value):
    """Convert a datetime or date value to a date object, or return None."""
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    return None


def format_date(value):
    """Format a datetime/date value as YYYY-MM-DD string, or return empty string."""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%Y-%m-%d")
    return str(value) if value else ""


def load_sheet_as_dict(wb, sheet_name, key_column):
    """Load a sheet into a dict keyed by the given column header."""
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    key_idx = headers.index(key_column)
    result = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[key_idx] is not None:
            result[row[key_idx]] = dict(zip(headers, row))
    return result


def load_bridge_table(wb, sheet_name):
    """Load a bridge table as a list of dicts."""
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))
    return rows


def main():
    parser = argparse.ArgumentParser(
        description="Export Scenario-Module overview from the FinOps Module Repository Excel file."
    )
    parser.add_argument(
        "--input", "-i",
        help="Path to the FinOpsCost-Module-Repository.xlsx file. "
             "If omitted, you will be prompted to enter it."
    )
    parser.add_argument(
        "--output", "-o",
        help="Output folder for the generated scenario_modules.csv. "
             "If omitted, you will be prompted to enter it."
    )
    parser.add_argument(
        "--no-changelog",
        action="store_true",
        default=False,
        help="Disable generation of the module changes Markdown file (enabled by default)."
    )
    parser.add_argument(
        "--days",
        type=int,
        default=30,
        help="Number of days to look back for module changes (default: 30)."
    )
    args = parser.parse_args()

    # Resolve input file
    excel_file = args.input
    if not excel_file:
        default_display = DEFAULT_EXCEL_FILE
        excel_file = input(f"Path to Excel file [{default_display}]: ").strip()
        if not excel_file:
            excel_file = DEFAULT_EXCEL_FILE

    if not os.path.isfile(excel_file):
        print(f"Error: File not found: {excel_file}", file=sys.stderr)
        sys.exit(1)

    # Resolve output folder
    output_dir = args.output
    if not output_dir:
        default_output = SCRIPT_DIR
        output_dir = input(f"Output folder [{default_output}]: ").strip()
        if not output_dir:
            output_dir = default_output

    if not os.path.isdir(output_dir):
        print(f"Error: Output folder not found: {output_dir}", file=sys.stderr)
        sys.exit(1)

    output_csv = os.path.join(output_dir, OUTPUT_FILENAME)

    wb = openpyxl.load_workbook(excel_file, data_only=True)

    # Entity lookups
    vbds = load_sheet_as_dict(wb, "VBD", "VBD Id")
    modules = load_sheet_as_dict(wb, "VBD Module", "Module Id")

    # Bridge tables
    map_vbd_scenario = load_bridge_table(wb, "Map-VBD-Scenario")
    map_scenario_module = load_bridge_table(wb, "Map-Scenario-Module")

    # Build join: VBD -> Scenario -> Module
    output_rows = []
    for vs in map_vbd_scenario:
        vbd_id = vs["VBD Id"]
        scenario_id = vs["Scenario Id"]
        vbd_title = vbds[vbd_id]["VBD Title"] if vbd_id in vbds else vs.get("VBD", "")

        for sm in map_scenario_module:
            if sm["Scenario Id"] != scenario_id:
                continue
            module_id = sm["Module Id"]
            module_title = modules[module_id]["Module Title"] if module_id in modules else sm.get("Module", "")
            scenario_title = sm.get("Scenario", vs.get("Scenario", ""))

            # Retrieve module date fields from entity lookup
            mod = modules.get(module_id, {})
            first_release = format_date(mod.get("Module first release date", ""))
            last_modified = format_date(mod.get("Module last modified", ""))
            lifecycle = mod.get("Module Lifecycle", "")

            output_rows.append({
                "VBD Title": vbd_title,
                "Scenario Title": scenario_title,
                "Module Id": module_id,
                "Module Title": module_title,
                "Suggested Order": sm.get("Suggested Order", ""),
                "Module Type": sm.get("Module Type", ""),
                "Module Lifecycle": lifecycle or "",
                "Module first release date": first_release or "",
                "Module last modified": last_modified or "",
            })

    # Sort by VBD Title, Scenario Title, Suggested Order
    output_rows.sort(key=lambda r: (
        r["VBD Title"],
        r["Scenario Title"],
        int(r["Suggested Order"]) if str(r["Suggested Order"]).isdigit() else 9999,
    ))

    # Write CSV
    fieldnames = ["VBD Title", "Scenario Title", "Module Id", "Module Title", "Suggested Order", "Module Type",
                  "Module Lifecycle", "Module first release date", "Module last modified"]
    with open(output_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(output_rows)

    print(f"Exported {len(output_rows)} rows to {output_csv}")

    # Generate module changes Markdown file
    if not args.no_changelog:
        cutoff = datetime.date.today() - datetime.timedelta(days=args.days)
        changed = []
        for mod in modules.values():
            first_release = to_date(mod.get("Module first release date"))
            last_modified = to_date(mod.get("Module last modified"))
            is_new = first_release is not None and first_release >= cutoff
            is_modified = (not is_new) and last_modified is not None and last_modified >= cutoff
            if is_new or is_modified:
                changed.append({
                    "Status": "New" if is_new else "Modified",
                    "Module Id": mod.get("Module Id", ""),
                    "Module Title": mod.get("Module Title", ""),
                    "Module first release date": format_date(first_release),
                    "Module last modified": format_date(last_modified),
                })

        # Sort: New first, then Modified; within each group by last modified descending
        changed.sort(key=lambda r: (0 if r["Status"] == "New" else 1, r["Module last modified"]), reverse=False)
        # Two-pass stable sort: first by date desc, then by status group
        changed.sort(key=lambda r: r["Module last modified"], reverse=True)
        changed.sort(key=lambda r: 0 if r["Status"] == "New" else 1)

        changelog_path = os.path.join(output_dir, CHANGELOG_FILENAME)
        today_str = datetime.date.today().strftime("%Y-%m-%d")
        with open(changelog_path, "w", encoding="utf-8") as f:
            f.write(f"# Module Changes — Last {args.days} Days\n\n")
            f.write(f"> **Last modified:** {today_str}\n\n")
            if changed:
                f.write("| Status | Module Id | Module Name | Module first release date | Module last modified |\n")
                f.write("|---|---|---|---|---|\n")
                for r in changed:
                    f.write(f"| {r['Status']} | {r['Module Id']} | {r['Module Title']} "
                            f"| {r['Module first release date']} | {r['Module last modified']} |\n")
            else:
                f.write("No module changes in the selected period.\n")

        print(f"Exported {len(changed)} module changes to {changelog_path}")


if __name__ == "__main__":
    main()
